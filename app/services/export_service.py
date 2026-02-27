import io
import os
from datetime import datetime, timezone
from sqlalchemy.orm import Session
from sqlalchemy import select
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from app.models.item import Item
from app.models.assignment import Assignment
from app.models.location import Location
from app.models.disposal import Disposal


# ── Font registration (česká diakritika) ──────────────────────────────────────
_FONT_REGULAR = "Helvetica"
_FONT_BOLD = "Helvetica-Bold"
_CZECH_CAPABLE = False

_FONT_PAIRS = [
    (
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
        "ESDejaVu", "ESDejaVuBold",
    ),
    (
        "/usr/share/fonts/ttf-dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/ttf-dejavu/DejaVuSans-Bold.ttf",
        "ESDejaVu", "ESDejaVuBold",
    ),
    (
        "/mnt/c/Windows/Fonts/arial.ttf",
        "/mnt/c/Windows/Fonts/arialbd.ttf",
        "ESArial", "ESArialBold",
    ),
]


def _init_export_fonts() -> None:
    global _FONT_REGULAR, _FONT_BOLD, _CZECH_CAPABLE
    if _CZECH_CAPABLE:
        return
    for reg_path, bold_path, reg_name, bold_name in _FONT_PAIRS:
        if not os.path.exists(reg_path):
            continue
        try:
            pdfmetrics.registerFont(TTFont(reg_name, reg_path))
            _FONT_REGULAR = reg_name
            if os.path.exists(bold_path):
                pdfmetrics.registerFont(TTFont(bold_name, bold_path))
                _FONT_BOLD = bold_name
            else:
                _FONT_BOLD = reg_name
            _CZECH_CAPABLE = True
            break
        except Exception:
            continue


_init_export_fonts()

_REASON_LABELS = {
    "liquidation": "Likvidace",
    "sale": "Prodej",
    "donation": "Darování",
    "theft": "Krádež",
    "loss": "Ztráta",
    "transfer": "Převod",
}


def export_items_excel(db: Session) -> bytes:
    from sqlalchemy import func as sqlfunc

    wb = Workbook()
    ws = wb.active
    ws.title = "Majetek"

    header_fill = PatternFill(start_color="1C2D42", end_color="1C2D42", fill_type="solid")
    header_font = Font(bold=True, color="F5A623", size=11)

    headers = ["ID", "Kód", "Název", "Kategorie", "S/N", "Zodpovědná osoba",
               "Datum nákupu", "Cena (Kč)", "Kód místnosti", "Název místnosti", "Aktivní", "Vytvořeno"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Pre-fetch current location per item (latest assignment)
    subq = (
        select(Assignment.item_id, sqlfunc.max(Assignment.assigned_at).label("max_at"))
        .group_by(Assignment.item_id)
        .subquery()
    )
    loc_rows = db.execute(
        select(Assignment.item_id, Location.code, Location.name)
        .join(subq, (Assignment.item_id == subq.c.item_id) & (Assignment.assigned_at == subq.c.max_at))
        .join(Location, Location.id == Assignment.location_id)
    ).all()
    item_loc = {row[0]: (row[1], row[2]) for row in loc_rows}  # item_id → (code, name)

    items = db.scalars(select(Item).order_by(Item.code)).all()
    for row_num, item in enumerate(items, 2):
        loc_code, loc_name = item_loc.get(item.id, ("", ""))
        ws.cell(row=row_num, column=1, value=item.id)
        ws.cell(row=row_num, column=2, value=item.code)
        ws.cell(row=row_num, column=3, value=item.name)
        ws.cell(row=row_num, column=4, value=item.category or "")
        ws.cell(row=row_num, column=5, value=item.serial_number or "")
        ws.cell(row=row_num, column=6, value=item.responsible_person or "")
        ws.cell(row=row_num, column=7, value=str(item.purchase_date) if item.purchase_date else "")
        ws.cell(row=row_num, column=8, value=float(item.purchase_price) if item.purchase_price else None)
        ws.cell(row=row_num, column=9, value=loc_code)
        ws.cell(row=row_num, column=10, value=loc_name)
        ws.cell(row=row_num, column=11, value="Ano" if item.is_active else "Ne")
        ws.cell(row=row_num, column=12, value=item.created_at.strftime("%d.%m.%Y"))

    col_widths = [6, 16, 32, 20, 20, 24, 14, 14, 16, 24, 8, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    # History sheet
    ws2 = wb.create_sheet("Historie přesunů")
    h2 = ["Kód položky", "Název položky", "Kód místnosti", "Název místnosti", "Přesunuto"]
    for col, h in enumerate(h2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
    ws2.freeze_panes = "A2"

    assignments = db.scalars(select(Assignment).order_by(Assignment.assigned_at)).all()
    for row_num, a in enumerate(assignments, 2):
        item = db.get(Item, a.item_id)
        loc = db.get(Location, a.location_id)
        ws2.cell(row=row_num, column=1, value=item.code if item else "")
        ws2.cell(row=row_num, column=2, value=item.name if item else "")
        ws2.cell(row=row_num, column=3, value=loc.code if loc else "")
        ws2.cell(row=row_num, column=4, value=loc.name if loc else "")
        ws2.cell(row=row_num, column=5, value=a.assigned_at.strftime("%d.%m.%Y %H:%M"))

    for i, w in enumerate([16, 32, 16, 24, 18], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_audit_pdf(db: Session, audit_id: int) -> bytes:
    report = _build_audit_report(db, audit_id)
    audit = report["audit"]

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    c.setTitle(f"Inventura: {audit.name}")
    pw, ph = A4
    margin = 20 * mm
    content_w = pw - 2 * margin
    header_h = 28 * mm
    footer_h = 12 * mm
    top_y = ph - header_h - 6 * mm   # první řádek obsahu pod hlavičkou
    bottom_y = footer_h + 4 * mm     # nejnižší dovolená pozice obsahu

    # Stav stránky
    page_num = [1]

    def _draw_header():
        # Oranžový pruh
        c.setFillColor(colors.HexColor("#F07800"))
        c.rect(0, ph - header_h, pw, header_h, fill=True, stroke=False)
        # Název inventury
        c.setFillColor(colors.white)
        c.setFont(_FONT_BOLD, 14)
        c.drawString(margin, ph - 10 * mm, f"Inventura: {audit.name}")
        # Metadata řádek
        c.setFont(_FONT_REGULAR, 8)
        created_name = audit.created_by_user.username if audit.created_by_user else "—"
        closed_name = audit.closed_by_user.username if audit.closed_by_user else "—"
        closed_at_str = audit.closed_at.strftime("%d.%m.%Y %H:%M") if audit.closed_at else "—"
        meta = (f"Zahájeno: {audit.started_at.strftime('%d.%m.%Y %H:%M')}  |  "
                f"Vytvořil: {created_name}  |  "
                f"Uzavřeno: {closed_at_str}  |  "
                f"Uzavřel: {closed_name}  |  "
                f"Stav: {'UZAVŘENA' if audit.status == 'closed' else 'OTEVŘENA'}")
        c.drawString(margin, ph - 18 * mm, meta[:120])
        # Statistiky
        c.setFont(_FONT_BOLD, 8)
        moved_part = f"    Přesunuto: {report['moved_count']}" if report.get("moved_count") else ""
        stats = (f"Naskenováno: {report['scanned_count']} / {report['total_items']}    "
                 f"Chybějící: {report['missing_count']}{moved_part}")
        c.drawString(margin, ph - 24 * mm, stats)

    def _draw_footer(page_n: int):
        c.setFillColor(colors.HexColor("#f0f0f0"))
        c.rect(0, 0, pw, footer_h, fill=True, stroke=False)
        c.setFillColor(colors.HexColor("#888888"))
        c.setFont(_FONT_REGULAR, 7)
        c.drawString(margin, 4 * mm, "AssetTrack — Interní inventarizační systém")
        c.drawRightString(pw - margin, 4 * mm,
                          f"Strana {page_n}  |  "
                          f"Vytištěno: {datetime.now(timezone.utc).strftime('%d.%m.%Y %H:%M')} UTC")

    def _new_page():
        _draw_footer(page_num[0])
        c.showPage()
        page_num[0] += 1
        _draw_header()
        return top_y

    # ── První stránka ──────────────────────────────────────────────────────────
    _draw_header()
    y = top_y

    # ── Skenované položky seskupené podle místností ────────────────────────────
    # Sestavit skupiny: {loc_name: [(item, scan, was_moved, from_loc_name), ...]}
    from collections import defaultdict
    groups: dict[str, list] = defaultdict(list)
    # scan_details obsahuje příznak was_moved; fallback pro případ starších volání bez tohoto pole
    for detail in report.get("scan_details", [{"scan": s, "was_moved": False, "from_location_name": None} for s in report["scans"]]):
        scan = detail["scan"]
        item = db.get(Item, scan.item_id)
        loc = db.get(Location, scan.location_id) if scan.location_id else None
        loc_name = loc.name if loc else "Neznámá místnost"
        groups[loc_name].append((item, scan, detail["was_moved"], detail["from_location_name"]))

    # Záhlaví sekce
    c.setFillColor(colors.HexColor("#333333"))
    c.rect(margin, y - 6 * mm, content_w, 6.5 * mm, fill=True, stroke=False)
    c.setFillColor(colors.white)
    c.setFont(_FONT_BOLD, 9)
    c.drawString(margin + 2 * mm, y - 3.5 * mm, "NASKENOVANÉ POLOŽKY")
    c.setFillColor(colors.black)
    y -= 10 * mm

    for loc_name in sorted(groups.keys()):
        items_in_loc = groups[loc_name]
        needed = 8 * mm + len(items_in_loc) * 6 * mm
        if y - needed < bottom_y:
            y = _new_page()

        # Záhlaví místnosti
        c.setFillColor(colors.HexColor("#F07800"))
        c.rect(margin, y - 5.5 * mm, content_w, 6 * mm, fill=True, stroke=False)
        c.setFillColor(colors.white)
        c.setFont(_FONT_BOLD, 9)
        c.drawString(margin + 3 * mm, y - 3 * mm, f"{loc_name}  ({len(items_in_loc)} ks)")
        c.setFillColor(colors.black)
        y -= 8 * mm

        for i, (item, scan, was_moved, from_loc_name) in enumerate(items_in_loc):
            row_h = 10 * mm if was_moved else 6 * mm
            if y - row_h < bottom_y:
                y = _new_page()
            bg = colors.HexColor("#fafafa") if i % 2 == 0 else colors.white
            c.setFillColor(bg)
            c.rect(margin, y - row_h + 0.5 * mm, content_w, row_h, fill=True, stroke=False)
            # Kód
            c.setFillColor(colors.black)
            c.setFont(_FONT_BOLD, 8)
            code = item.code if item else "?"
            c.drawString(margin + 2 * mm, y - 3 * mm, code)
            # Název
            c.setFont(_FONT_REGULAR, 8)
            name_max = 52 if was_moved else 60
            name = (item.name if item else "?")[:name_max]
            c.drawString(margin + 30 * mm, y - 3 * mm, name)
            # Čas skenu
            scanned_at = scan.scanned_at.strftime("%d.%m.%Y %H:%M") if scan.scanned_at else ""
            c.setFont(_FONT_REGULAR, 7)
            c.setFillColor(colors.HexColor("#666666"))
            c.drawRightString(pw - margin - 2 * mm, y - 3 * mm, scanned_at)
            # Přesun — druhý řádek
            if was_moved and from_loc_name:
                c.setFont(_FONT_REGULAR, 7)
                c.setFillColor(colors.HexColor("#C05000"))
                c.drawString(margin + 30 * mm, y - 7.5 * mm,
                             f"\u2192 přesunuto z: {from_loc_name}")
            c.setFillColor(colors.black)
            y -= row_h

        y -= 3 * mm  # mezera mezi skupinami

    # ── Chybějící položky ─────────────────────────────────────────────────────
    if report["missing_items"]:
        if y - 14 * mm < bottom_y:
            y = _new_page()

        y -= 4 * mm
        c.setFillColor(colors.HexColor("#C00000"))
        c.rect(margin, y - 6 * mm, content_w, 6.5 * mm, fill=True, stroke=False)
        c.setFillColor(colors.white)
        c.setFont(_FONT_BOLD, 9)
        c.drawString(margin + 2 * mm, y - 3.5 * mm,
                     f"NENASKENOVANÉ POLOŽKY  ({len(report['missing_items'])} ks)")
        c.setFillColor(colors.black)
        y -= 10 * mm

        for i, item in enumerate(report["missing_items"]):
            if y < bottom_y:
                y = _new_page()
            bg = colors.HexColor("#fff5f5") if i % 2 == 0 else colors.white
            c.setFillColor(bg)
            c.rect(margin, y - 5 * mm, content_w, 5.5 * mm, fill=True, stroke=False)
            c.setFillColor(colors.HexColor("#C00000"))
            c.setFont(_FONT_BOLD, 8)
            c.drawString(margin + 2 * mm, y - 3 * mm, item.code)
            c.setFillColor(colors.black)
            c.setFont(_FONT_REGULAR, 8)
            c.drawString(margin + 30 * mm, y - 3 * mm, (item.name or "")[:60])
            y -= 6 * mm

    # ── Poslední stránka — zápatí ──────────────────────────────────────────────
    _draw_footer(page_num[0])
    c.save()
    return buf.getvalue()


def export_disposals_excel(db: Session) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Vyřazený majetek"

    # Záhlaví
    headers = ["ID", "Kód položky", "Název položky", "Kategorie", "S/N", "Pořizovací cena",
               "Důvod vyřazení", "Datum vyřazení", "Číslo dokladu", "Poznámka"]
    header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data
    disposals = db.scalars(select(Disposal).order_by(Disposal.disposed_at.desc())).all()
    for row_num, d in enumerate(disposals, 2):
        item = db.get(Item, d.item_id)
        ws.cell(row=row_num, column=1, value=d.id)
        ws.cell(row=row_num, column=2, value=item.code if item else "")
        ws.cell(row=row_num, column=3, value=item.name if item else "")
        ws.cell(row=row_num, column=4, value=item.category if item else "")
        ws.cell(row=row_num, column=5, value=item.serial_number if item else "")
        ws.cell(row=row_num, column=6, value=float(item.purchase_price) if item and item.purchase_price else None)
        ws.cell(row=row_num, column=7, value=_REASON_LABELS.get(d.reason, d.reason))
        ws.cell(row=row_num, column=8, value=d.disposed_at.strftime("%d.%m.%Y %H:%M") if d.disposed_at else "")
        ws.cell(row=row_num, column=9, value=d.document_ref or "")
        ws.cell(row=row_num, column=10, value=d.note or "")

    # Šířky sloupců
    col_widths = [6, 18, 30, 15, 20, 16, 15, 20, 20, 40]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_disposal_pdf(db: Session, disposal_id: int) -> bytes:
    disposal = db.get(Disposal, disposal_id)
    if not disposal:
        from fastapi import HTTPException
        raise HTTPException(status_code=404, detail="Záznam o vyřazení nenalezen")

    item = db.get(Item, disposal.item_id)
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    pw, ph = A4
    margin = 20 * mm

    # ── Hlavička ──────────────────────────────────────────────────────────────
    c.setFillColor(colors.HexColor("#C00000"))
    c.rect(0, ph - 35 * mm, pw, 35 * mm, fill=True, stroke=False)
    c.setFillColor(colors.white)
    c.setFont(_FONT_BOLD, 18)
    c.drawString(margin, ph - 20 * mm, "PROTOKOL O VYŘAZENÍ MAJETKU")
    c.setFont(_FONT_REGULAR, 10)
    c.drawString(margin, ph - 29 * mm, f"AssetTrack  \u00b7  Číslo záznamu: {disposal.id}")

    # ── Datum tisku ────────────────────────────────────────────────────────────
    c.setFillColor(colors.black)
    c.setFont(_FONT_REGULAR, 9)
    c.drawRightString(pw - margin, ph - 40 * mm,
                      f"Vytištěno: {datetime.now(timezone.utc).strftime('%d.%m.%Y %H:%M')} UTC")

    # ── Sekce: Identifikace položky ────────────────────────────────────────────
    y = ph - 50 * mm
    y = _pdf_section_header(c, "Identifikace majetku", y, margin, pw)
    rows_item = [
        ("Kód položky:", item.code if item else "—"),
        ("Název:", item.name if item else "—"),
        ("Kategorie:", item.category or "—" if item else "—"),
        ("Sériové číslo:", item.serial_number or "—" if item else "—"),
        ("Zodpovědná osoba:", item.responsible_person or "—" if item else "—"),
        ("Datum nákupu:", str(item.purchase_date) if item and item.purchase_date else "—"),
        ("Pořizovací cena:", f"{item.purchase_price:,.2f} Kč" if item and item.purchase_price else "—"),
    ]
    y = _pdf_table(c, rows_item, y, margin, pw)

    # ── Sekce: Údaje o vyřazení ────────────────────────────────────────────────
    y -= 8 * mm
    y = _pdf_section_header(c, "Údaje o vyřazení", y, margin, pw)
    rows_disposal = [
        ("Důvod vyřazení:", _REASON_LABELS.get(disposal.reason, disposal.reason)),
        ("Datum vyřazení:", disposal.disposed_at.strftime("%d.%m.%Y %H:%M") if disposal.disposed_at else "—"),
        ("Číslo dokladu:", disposal.document_ref or "—"),
        ("Poznámka:", disposal.note or "—"),
    ]
    y = _pdf_table(c, rows_disposal, y, margin, pw)

    # ── Podpisové řádky ────────────────────────────────────────────────────────
    y -= 20 * mm
    sig_y = y
    sig_w = (pw - 2 * margin - 20 * mm) / 2

    # Levý sloupec
    c.setFont(_FONT_REGULAR, 10)
    c.drawString(margin, sig_y, "Vyřadil / Vyřadila:")
    c.line(margin, sig_y - 12 * mm, margin + sig_w, sig_y - 12 * mm)
    c.setFont(_FONT_REGULAR, 8)
    c.setFillColor(colors.gray)
    c.drawString(margin, sig_y - 15 * mm, "Jméno, podpis, datum")

    # Pravý sloupec
    right_x = pw - margin - sig_w
    c.setFillColor(colors.black)
    c.setFont(_FONT_REGULAR, 10)
    c.drawString(right_x, sig_y, "Schválil / Schválila:")
    c.line(right_x, sig_y - 12 * mm, right_x + sig_w, sig_y - 12 * mm)
    c.setFont(_FONT_REGULAR, 8)
    c.setFillColor(colors.gray)
    c.drawString(right_x, sig_y - 15 * mm, "Jméno, podpis, datum")

    # ── Zápatí ────────────────────────────────────────────────────────────────
    c.setFillColor(colors.HexColor("#f0f0f0"))
    c.rect(0, 0, pw, 12 * mm, fill=True, stroke=False)
    c.setFillColor(colors.gray)
    c.setFont(_FONT_REGULAR, 8)
    c.drawString(margin, 4 * mm, "AssetTrack \u2014 Interní inventarizační systém")
    c.drawRightString(pw - margin, 4 * mm, f"Protokol č. {disposal.id}")

    c.save()
    return buf.getvalue()


def _pdf_section_header(c, title: str, y: float, margin: float, pw: float) -> float:
    c.setFillColor(colors.HexColor("#404040"))
    c.rect(margin, y - 6 * mm, pw - 2 * margin, 7 * mm, fill=True, stroke=False)
    c.setFillColor(colors.white)
    c.setFont(_FONT_BOLD, 10)
    c.drawString(margin + 3 * mm, y - 3.5 * mm, title)
    c.setFillColor(colors.black)
    return y - 10 * mm


def _pdf_table(c, rows: list[tuple], y: float, margin: float, pw: float) -> float:
    col1_w = 55 * mm
    for i, (label, value) in enumerate(rows):
        bg = colors.HexColor("#f8f8f8") if i % 2 == 0 else colors.white
        c.setFillColor(bg)
        c.rect(margin, y - 6 * mm, pw - 2 * margin, 7 * mm, fill=True, stroke=False)
        c.setFillColor(colors.black)
        c.setFont(_FONT_BOLD, 9)
        c.drawString(margin + 2 * mm, y - 3 * mm, label)
        c.setFont(_FONT_REGULAR, 9)
        c.drawString(margin + col1_w, y - 3 * mm, str(value)[:70])
        y -= 7 * mm
    return y


def _build_audit_report(db: Session, audit_id: int) -> dict:
    from app.services.audit_service import get_audit_report
    return get_audit_report(db, audit_id)
