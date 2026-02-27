"""
Import service — hromadný import majetku z Excel souboru.
Podporuje formáty .xlsx a .xlsm (openpyxl).
"""
import io
from datetime import datetime, date
from decimal import Decimal, InvalidOperation
from sqlalchemy.orm import Session
from sqlalchemy import select, func
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app.models.item import Item
from app.models.location import Location
from app.models.assignment import Assignment

# Mapování názvů sloupců (malá písmena, bez diakritiky) → interní název pole
_COL_MAP = {
    # Kód
    "kód": "code", "kod": "code", "code": "code",
    # Název
    "název": "name", "nazev": "name", "name": "name",
    # Kategorie
    "kategorie": "category", "category": "category",
    # Popis
    "popis": "description", "description": "description",
    # Sériové číslo
    "sériové číslo": "serial_number", "seriove cislo": "serial_number",
    "s/n": "serial_number", "sn": "serial_number", "serial_number": "serial_number",
    "sériové č.": "serial_number", "ser. číslo": "serial_number",
    # Datum nákupu
    "datum nákupu": "purchase_date", "datum nakupu": "purchase_date",
    "purchase_date": "purchase_date", "datum": "purchase_date",
    # Cena pořízení
    "cena pořízení (kč)": "purchase_price", "cena pořízení": "purchase_price",
    "cena porizeni": "purchase_price", "cena": "purchase_price",
    "purchase_price": "purchase_price", "cena (kč)": "purchase_price",
    # Zodpovědná osoba
    "zodpovědná osoba": "responsible_person", "zodpovedna osoba": "responsible_person",
    "zodpovědný": "responsible_person", "odpovědná osoba": "responsible_person",
    "responsible_person": "responsible_person", "responsible": "responsible_person",
    # Kód lokace
    "kód lokace": "location_code", "kod lokace": "location_code",
    "lokace": "location_code", "location_code": "location_code",
    "location": "location_code",
}


def _normalize(s: str) -> str:
    """Normalizuje text pro porovnání hlaviček (lowercase, strip, bez hvězdičky).
    Hvězdička (*) v šabloně označuje povinné pole — při detekci hlavičky ji ignorujeme.
    """
    return str(s).lower().strip().rstrip("*").strip()


def _parse_date(value) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    if not s:
        return None
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d.%m.%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_price(value) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    s = str(value).replace("\xa0", "").replace(" ", "").replace(",", ".").replace("Kč", "").replace("kč", "").strip()
    if not s:
        return None
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def _generate_code(db: Session, used: set) -> str:
    """Vygeneruje jedinečný kód ve formátu IT-NNNNN."""
    count = db.scalar(select(func.count()).select_from(Item)) or 0
    while True:
        code = f"IT-{count + 1:05d}"
        if code not in used:
            existing = db.scalar(select(Item).where(Item.code == code))
            if not existing:
                return code
        count += 1


_IMPORT_MAX_ROWS = 2000


def import_items_from_excel(db: Session, file_data: bytes) -> dict:
    """
    Zpracuje Excel soubor a importuje položky majetku.

    Vrací dict s:
      - success: bool
      - imported: int
      - skipped: int
      - errors: int
      - details: list[dict] — výsledek pro každou řádku
      - error: str (jen pokud success=False)
    """
    try:
        wb = load_workbook(filename=io.BytesIO(file_data), data_only=True)
    except Exception as e:
        return {"success": False, "error": f"Nepodařilo se načíst soubor: {e}", "imported": 0, "skipped": 0, "errors": 0, "details": []}

    # Hledáme první list (nebo list pojmenovaný "Import majetku")
    if "Import majetku" in wb.sheetnames:
        ws = wb["Import majetku"]
    else:
        ws = wb.active

    # Najdeme řádek s hlavičkami (prohledáme prvních 10 řádků)
    header_row = None
    col_map: dict[str, int] = {}  # field_name → column index (1-based)

    for row in ws.iter_rows(min_row=1, max_row=10):
        matched = 0
        tmp_map: dict[str, int] = {}
        for cell in row:
            if cell.value is None:
                continue
            key = _normalize(str(cell.value))
            if key in _COL_MAP:
                tmp_map[_COL_MAP[key]] = cell.column
                matched += 1
        if matched >= 1 and "name" in tmp_map:
            header_row = row[0].row
            col_map = tmp_map
            break

    if header_row is None or "name" not in col_map:
        return {
            "success": False,
            "error": (
                "Nepodařilo se najít povinný sloupec 'Název' v souboru. "
                "Zkontrolujte, zda jste použili šablonu AssetTrack."
            ),
            "imported": 0, "skipped": 0, "errors": 0, "details": [],
        }

    def get_val(row_values: tuple, field: str):
        idx = col_map.get(field)
        if idx is None:
            return None
        v = row_values[idx - 1]
        if v is None:
            return None
        return str(v).strip() if not isinstance(v, (int, float, date, datetime)) else v

    # --- Kontrola limitu řádků ---
    data_rows = ws.max_row - header_row
    if data_rows > _IMPORT_MAX_ROWS:
        return {
            "success": False,
            "error": f"Soubor obsahuje příliš mnoho řádků ({data_rows}). Maximum je {_IMPORT_MAX_ROWS}.",
            "imported": 0, "skipped": 0, "errors": 0, "details": [],
        }

    # --- Fáze 1: validace a příprava dat (bez zápisu do DB) ---
    to_insert: list[dict] = []
    results: list[dict] = []
    used_codes: set[str] = set()

    for row_values in ws.iter_rows(min_row=header_row + 1, max_row=header_row + _IMPORT_MAX_ROWS, values_only=True):
        # Přeskočíme prázdné řádky
        if all(v is None or (isinstance(v, str) and v.strip() == "") for v in row_values):
            continue

        name_raw = get_val(row_values, "name")
        name = str(name_raw).strip() if name_raw else None
        if not name:
            results.append({"status": "skipped", "code": "—", "name": "—", "reason": "Chybí název (povinný sloupec)"})
            continue

        # Kód
        code_raw = get_val(row_values, "code")
        code = str(code_raw).strip() if code_raw else None

        if code:
            # Kontrola duplikátu v DB
            if code in used_codes or db.scalar(select(Item).where(Item.code == code)):
                results.append({"status": "skipped", "code": code, "name": name, "reason": f"Kód '{code}' již existuje nebo se opakuje v souboru"})
                continue
        else:
            code = _generate_code(db, used_codes)

        used_codes.add(code)

        # Datum a cena
        purchase_date = _parse_date(get_val(row_values, "purchase_date"))
        purchase_price = _parse_price(get_val(row_values, "purchase_price"))

        # Lokace (jen validace — přiřadíme ve fázi 2)
        location_code_raw = get_val(row_values, "location_code")
        location_code = str(location_code_raw).strip() if location_code_raw else None
        location_id = None
        location_note = None

        if location_code:
            loc = db.scalar(select(Location).where(Location.code == location_code, Location.is_active == True))
            if loc:
                location_id = loc.id
            else:
                location_note = f"Lokace '{location_code}' nenalezena — položka importována bez přiřazení"

        to_insert.append({
            "code": code,
            "name": name,
            "category": get_val(row_values, "category") or None,
            "description": get_val(row_values, "description") or None,
            "serial_number": get_val(row_values, "serial_number") or None,
            "responsible_person": get_val(row_values, "responsible_person") or None,
            "purchase_date": purchase_date,
            "purchase_price": purchase_price,
            "location_id": location_id,
            "location_note": location_note,
        })

    # --- Fáze 2: zápis do DB ---
    imported = 0
    errors = 0

    for d in to_insert:
        try:
            item = Item(
                code=d["code"],
                name=d["name"],
                category=d["category"],
                description=d["description"],
                serial_number=d["serial_number"],
                responsible_person=d["responsible_person"],
                purchase_date=d["purchase_date"],
                purchase_price=d["purchase_price"],
            )
            db.add(item)
            db.flush()  # získáme item.id

            if d["location_id"]:
                db.add(Assignment(item_id=item.id, location_id=d["location_id"]))

            imported += 1
            results.append({
                "status": "imported",
                "code": d["code"],
                "name": d["name"],
                "reason": d.get("location_note") or "",
            })
        except Exception as e:
            db.rollback()
            errors += 1
            results.append({
                "status": "error",
                "code": d.get("code", "—"),
                "name": d.get("name", "—"),
                "reason": str(e),
            })

    if imported > 0:
        try:
            db.commit()
        except Exception as e:
            db.rollback()
            return {
                "success": False,
                "error": f"Chyba při ukládání do databáze: {e}",
                "imported": 0, "skipped": len([r for r in results if r["status"] == "skipped"]),
                "errors": len(to_insert), "details": results,
            }

    skipped = len([r for r in results if r["status"] == "skipped"])

    return {
        "success": True,
        "imported": imported,
        "skipped": skipped,
        "errors": errors,
        "details": results,
    }


def generate_import_template() -> bytes:
    """Vygeneruje Excel šablonu pro import majetku."""
    wb = Workbook()

    # ── List 1: Data ──────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Import majetku"

    header_fill = PatternFill(start_color="1C2D42", end_color="1C2D42", fill_type="solid")
    header_font = Font(bold=True, color="F5A623", size=11)
    example_fill = PatternFill(start_color="EEF4FF", end_color="EEF4FF", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")

    headers = [
        "Kód",
        "Název *",
        "Kategorie",
        "Popis",
        "Sériové číslo",
        "Zodpovědná osoba",
        "Datum nákupu",
        "Cena pořízení (Kč)",
        "Kód lokace",
    ]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # Ukázkové řádky
    examples = [
        ["IT-00001", "Notebook Dell Latitude 5540", "Výpočetní technika", "Pracovní notebook – 15,6\"", "SN-ABC123456", "Jan Novák",  "15.01.2024", 34990, "A101"],
        ["",         "Monitor LG 27\" UHD 4K",       "Výpočetní technika", "",                           "SN-DEF789012", "Jan Novák",  "20.02.2024", 12500, "A101"],
        ["",         "Kancelářská židle Ergon Pro",   "Nábytek",           "Ergonomická, bederní opěrka", "",             "Jana Dvořák","",           8900,  "B201"],
        ["IT-99999", "Projektor Epson EB-W51",        "AV technika",       "Přenosný, WXGA rozlišení",    "SN-GHI345678", "",           "10.03.2023", 21000, ""],
    ]

    for row_num, ex in enumerate(examples, 2):
        for col, val in enumerate(ex, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.fill = example_fill

    # Šířky sloupců
    widths = [14, 36, 22, 34, 20, 24, 16, 20, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"

    # ── List 2: Pokyny ────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Pokyny k importu")
    ws2.column_dimensions["A"].width = 90

    bold_font = Font(bold=True, size=11)
    title_font = Font(bold=True, size=13)

    lines = [
        ("POKYNY K IMPORTU MAJETKU DO SYSTÉMU ASSETTRACK", title_font, 22),
        ("", None, 8),
        ("1. POSTUP IMPORTU", bold_font, 18),
        ("   a) Vyplňte tabulku v listu 'Import majetku' (tento soubor).", None, 15),
        ("   b) Šablona obsahuje ukázkové řádky (modré) — ty před importem smažte.", None, 15),
        ("   c) Uložte soubor jako .xlsx (Excel 2007 a novější).", None, 15),
        ("   d) V AssetTrack přejděte do sekce Import a nahrajte soubor.", None, 15),
        ("   e) Zkontrolujte výsledky importu.", None, 15),
        ("", None, 8),
        ("2. POVINNÉ SLOUPCE", bold_font, 18),
        ("   • Název — název vybavení musí být vyplněn u každé řádky.", None, 15),
        ("", None, 8),
        ("3. VOLITELNÉ SLOUPCE", bold_font, 18),
        ("   • Kód        — pokud je prázdný, systém vygeneruje kód ve formátu IT-NNNNN.", None, 15),
        ("   • Kategorie  — libovolný text (Výpočetní technika, Nábytek, AV technika…).", None, 15),
        ("   • Popis      — volný popis, nepovinný.", None, 15),
        ("   • Sériové č. — výrobní S/N z výrobního štítku.", None, 15),
        ("   • Datum nák. — formát DD.MM.RRRR nebo RRRR-MM-DD.", None, 15),
        ("   • Cena (Kč)  — číslo v Kč, desetinnou část oddělte tečkou (např. 12500.50).", None, 15),
        ("   • Kód lokace — musí odpovídat existujícímu kódu lokace v systému.", None, 15),
        ("", None, 8),
        ("4. PRAVIDLA A OMEZENÍ", bold_font, 18),
        ("   • Řádky s prázdným sloupcem Název jsou přeskočeny.", None, 15),
        ("   • Řádky s duplicitním Kódem jsou přeskočeny (kód musí být jedinečný).", None, 15),
        ("   • Pokud Kód lokace neexistuje, položka se importuje bez přiřazení k lokaci.", None, 15),
        ("   • Import NIKDY nepřepisuje existující data — přidává pouze nové položky.", None, 15),
        ("   • Maximální velikost souboru: 10 MB.", None, 15),
        ("", None, 8),
        ("5. PŘÍPUSTNÉ FORMÁTY SOUBORU", bold_font, 18),
        ("   • .xlsx (Excel 2007 a novější) — doporučeno", None, 15),
        ("   • .xlsm (sešit s makry — makra jsou ignorována)", None, 15),
    ]

    for row_num, (text, font, height) in enumerate(lines, 1):
        cell = ws2.cell(row=row_num, column=1, value=text)
        if font:
            cell.font = font
        ws2.row_dimensions[row_num].height = height

    wb.active = ws

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
