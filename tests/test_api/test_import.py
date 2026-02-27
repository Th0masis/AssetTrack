"""
Testy pro import majetku z Excel souboru.
- GET /api/import/template → stažení šablony
- GET /import → stránka importu
- POST /import → zpracování souboru
- import_service: validace a import dat
"""
import io
import pytest
from openpyxl import Workbook
from app.services import import_service as svc


# ── Pomocné funkce ────────────────────────────────────────────────────────────

def make_excel(rows: list[list], headers: list[str] | None = None) -> bytes:
    """Vytvoří Excel soubor v paměti s danými záhlavími a řádky."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Import majetku"

    if headers is None:
        headers = ["Kód", "Název *", "Kategorie", "Popis", "Sériové číslo",
                   "Datum nákupu", "Cena pořízení (Kč)", "Kód lokace"]

    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    for row_num, row in enumerate(rows, 2):
        for col, val in enumerate(row, 1):
            ws.cell(row=row_num, column=col, value=val)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Testy šablony (API) ───────────────────────────────────────────────────────

class TestImportTemplate:
    def test_template_download_returns_xlsx(self, client):
        res = client.get("/api/import/template")
        assert res.status_code == 200
        assert "spreadsheetml" in res.headers["content-type"]
        assert "attachment" in res.headers["content-disposition"]
        assert res.content  # non-empty

    def test_template_is_valid_excel(self, client):
        res = client.get("/api/import/template")
        wb = Workbook()
        # Ověříme, že obsah je čitelný jako Excel
        buf = io.BytesIO(res.content)
        from openpyxl import load_workbook
        wb = load_workbook(buf)
        assert "Import majetku" in wb.sheetnames
        assert "Pokyny k importu" in wb.sheetnames

    def test_template_has_required_headers(self, client):
        from openpyxl import load_workbook
        res = client.get("/api/import/template")
        wb = load_workbook(io.BytesIO(res.content))
        ws = wb["Import majetku"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, 9)]
        assert any("Název" in str(h) for h in headers if h)
        assert any("Kód" in str(h) for h in headers if h)


# ── Testy UI stránky ──────────────────────────────────────────────────────────

class TestImportPage:
    def test_get_import_page(self, client):
        res = client.get("/import")
        assert res.status_code == 200
        assert "Import" in res.text

    def test_import_page_shows_upload_form(self, client):
        res = client.get("/import")
        assert "dropzone" in res.text
        assert 'type="file"' in res.text

    def test_import_page_has_template_download_link(self, client):
        res = client.get("/import")
        assert "/api/import/template" in res.text


# ── Testy POST /import ────────────────────────────────────────────────────────

class TestImportPost:
    def test_import_valid_items(self, client):
        data = make_excel([
            ["", "Notebook Dell XPS", "Výpočetní technika", "Pracovní NB", "SN-001", "15.01.2024", "35000", ""],
            ["", "Monitor LG 27\"",   "Výpočetní technika", "",             "SN-002", "",           "12000", ""],
        ])
        res = client.post(
            "/import",
            files={"file": ("import.xlsx", data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert res.status_code == 200
        assert "Importováno" in res.text

    def test_import_rejects_non_excel(self, client):
        res = client.post(
            "/import",
            files={"file": ("data.csv", b"col1,col2\nval1,val2", "text/csv")},
        )
        assert res.status_code == 200
        assert "Nepodporovaný formát" in res.text

    def test_import_shows_results_table(self, client):
        data = make_excel([
            ["", "Testovací položka", "Kategorie", "", "", "", "", ""],
        ])
        res = client.post(
            "/import",
            files={"file": ("import.xlsx", data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert res.status_code == 200
        assert "Testovací položka" in res.text


# ── Testy import service ──────────────────────────────────────────────────────

class TestImportService:
    def test_import_basic_item(self, client):
        from sqlalchemy import create_engine
        from sqlalchemy.orm import sessionmaker
        from sqlalchemy.pool import StaticPool
        from app.database import Base
        from app.models.item import Item

        engine = create_engine("sqlite:///:memory:", connect_args={"check_same_thread": False}, poolclass=StaticPool)
        Base.metadata.create_all(engine)
        Session = sessionmaker(bind=engine)
        db = Session()

        data = make_excel([["", "Notebook Test", "IT", "", "SN-X", "", "", ""]])
        result = svc.import_items_from_excel(db, data)

        assert result["success"] is True
        assert result["imported"] == 1
        assert result["skipped"] == 0
        assert result["errors"] == 0

        items = db.query(Item).all()
        assert len(items) == 1
        assert items[0].name == "Notebook Test"
        assert items[0].code.startswith("IT-")

        db.close()

    def test_import_with_custom_code(self, client):
        from sqlalchemy import create_engine
        from sqlalchemy.orm import sessionmaker
        from sqlalchemy.pool import StaticPool
        from app.database import Base
        from app.models.item import Item

        engine = create_engine("sqlite:///:memory:", connect_args={"check_same_thread": False}, poolclass=StaticPool)
        Base.metadata.create_all(engine)
        Session = sessionmaker(bind=engine)
        db = Session()

        data = make_excel([["CUSTOM-001", "Počítač", "IT", "", "", "", "", ""]])
        result = svc.import_items_from_excel(db, data)

        assert result["success"] is True
        assert result["imported"] == 1
        item = db.query(Item).first()
        assert item.code == "CUSTOM-001"

        db.close()

    def test_import_skips_empty_name(self, client):
        from sqlalchemy import create_engine
        from sqlalchemy.orm import sessionmaker
        from sqlalchemy.pool import StaticPool
        from app.database import Base

        engine = create_engine("sqlite:///:memory:", connect_args={"check_same_thread": False}, poolclass=StaticPool)
        Base.metadata.create_all(engine)
        Session = sessionmaker(bind=engine)
        db = Session()

        data = make_excel([
            ["", "",             "IT", "", "", "", "", ""],  # prázdný název → skip
            ["", "Validní název", "IT", "", "", "", "", ""],  # ok
        ])
        result = svc.import_items_from_excel(db, data)

        assert result["imported"] == 1
        assert result["skipped"] == 1

        db.close()

    def test_import_skips_duplicate_code(self, client):
        from sqlalchemy import create_engine
        from sqlalchemy.orm import sessionmaker
        from sqlalchemy.pool import StaticPool
        from app.database import Base
        from app.models.item import Item
        from datetime import datetime, timezone

        engine = create_engine("sqlite:///:memory:", connect_args={"check_same_thread": False}, poolclass=StaticPool)
        Base.metadata.create_all(engine)
        Session = sessionmaker(bind=engine)
        db = Session()

        # Předvytvořená položka s kódem DUPE-001
        existing = Item(code="DUPE-001", name="Existující")
        db.add(existing)
        db.commit()

        data = make_excel([
            ["DUPE-001", "Duplicitní", "", "", "", "", "", ""],
            ["",         "Nová",       "", "", "", "", "", ""],
        ])
        result = svc.import_items_from_excel(db, data)

        assert result["skipped"] == 1
        assert result["imported"] == 1

        db.close()

    def test_import_multiple_items_generate_unique_codes(self, client):
        from sqlalchemy import create_engine
        from sqlalchemy.orm import sessionmaker
        from sqlalchemy.pool import StaticPool
        from app.database import Base
        from app.models.item import Item

        engine = create_engine("sqlite:///:memory:", connect_args={"check_same_thread": False}, poolclass=StaticPool)
        Base.metadata.create_all(engine)
        Session = sessionmaker(bind=engine)
        db = Session()

        data = make_excel([
            ["", "Položka A", "", "", "", "", "", ""],
            ["", "Položka B", "", "", "", "", "", ""],
            ["", "Položka C", "", "", "", "", "", ""],
        ])
        result = svc.import_items_from_excel(db, data)

        assert result["imported"] == 3
        items = db.query(Item).all()
        codes = [i.code for i in items]
        assert len(set(codes)) == 3  # všechny kódy jsou unikátní

        db.close()

    def test_import_fails_without_name_column(self, client):
        from sqlalchemy import create_engine
        from sqlalchemy.orm import sessionmaker
        from sqlalchemy.pool import StaticPool
        from app.database import Base

        engine = create_engine("sqlite:///:memory:", connect_args={"check_same_thread": False}, poolclass=StaticPool)
        Base.metadata.create_all(engine)
        Session = sessionmaker(bind=engine)
        db = Session()

        # Excel bez sloupce Název
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NeexistujiciSloupec")
        ws.cell(row=2, column=1, value="data")
        buf = io.BytesIO()
        wb.save(buf)

        result = svc.import_items_from_excel(db, buf.getvalue())

        assert result["success"] is False
        assert "Název" in result["error"]

        db.close()

    def test_generate_import_template_returns_bytes(self):
        result = svc.generate_import_template()
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_template_content(self):
        from openpyxl import load_workbook
        data = svc.generate_import_template()
        wb = load_workbook(io.BytesIO(data))
        assert "Import majetku" in wb.sheetnames
        ws = wb["Import majetku"]
        # Alespoň záhlaví + jeden ukázkový řádek
        assert ws.max_row >= 2
